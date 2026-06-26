"""BI 数据源管理测试 — F-DS-01/02/03/04/05/07（API-BI-01/02/03）。

落盘目录通过 fixture 重定向到 tmp_path，避免污染真实 data/bi_datasets。
"""
import io

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app.main import app
from app.services import bi_service


@pytest.fixture(autouse=True)
def _tmp_data_dir(tmp_path, monkeypatch):
    """每个用例独立的落盘目录。"""
    monkeypatch.setattr(bi_service, "BI_DATA_DIR", tmp_path / "bi_datasets")


@pytest.fixture
def client():
    with TestClient(app) as c:
        yield c


def _csv_bytes(text: str) -> bytes:
    return text.encode("utf-8-sig")


def _upload(client, filename: str, content: bytes):
    return client.post(
        "/api/bi/datasets/upload",
        files={"file": (filename, content, "application/octet-stream")},
    )


# ─── F-DS-01 / F-DS-03：CSV 上传 + 类型推断 ──────────────────────
def test_upload_csv_infers_types_and_roles(client):
    csv = "地区,销售额\n华东,1200\n华南,980\n"
    resp = _upload(client, "sales.csv", _csv_bytes(csv))
    assert resp.status_code == 200
    body = resp.json()
    assert body["code"] == 0
    meta = body["data"]
    assert meta["rowCount"] == 2
    assert meta["sourceFilename"] == "sales.csv"
    cols = {c["name"]: c for c in meta["columns"]}
    assert cols["地区"]["type"] == "string"
    assert cols["地区"]["role"] == "dimension"
    assert cols["销售额"]["type"] == "number"
    assert cols["销售额"]["role"] == "measure"


# ─── F-DS-03：高唯一率整数列疑似 ID → 维度 ───────────────────────
def test_id_like_column_defaults_to_dimension(client):
    csv = "订单号,金额\n1001,50\n1002,60\n1003,70\n"
    meta = _upload(client, "orders.csv", _csv_bytes(csv)).json()["data"]
    cols = {c["name"]: c for c in meta["columns"]}
    assert cols["订单号"]["role"] == "dimension"
    assert cols["金额"]["role"] == "measure"


# ─── F-DS-02：Excel 多级表头拍平为 "上级/子级" ───────────────────
def test_upload_xlsx_multilevel_header_flatten(client):
    wb = openpyxl.Workbook()
    ws = wb.active
    # 第 1 行：地区(合并) | 销售(合并两列)
    ws["A1"] = "地区"
    ws["B1"] = "销售"
    ws.merge_cells("B1:C1")
    ws.merge_cells("A1:A2")
    # 第 2 行子表头
    ws["B2"] = "Q1"
    ws["C2"] = "Q2"
    ws["A3"] = "华东"
    ws["B3"] = 100
    ws["C3"] = 200
    buf = io.BytesIO()
    wb.save(buf)

    meta = _upload(client, "report.xlsx", buf.getvalue()).json()["data"]
    names = [c["name"] for c in meta["columns"]]
    assert "地区" in names
    assert "销售/Q1" in names
    assert "销售/Q2" in names
    assert meta["rowCount"] == 1


# ─── F-DS-05：列表回显 + F-DS-07：删除 ───────────────────────────
def test_list_and_delete_dataset(client):
    meta = _upload(client, "a.csv", _csv_bytes("x,y\n1,2\n")).json()["data"]
    dataset_id = meta["id"]

    listed = client.get("/api/bi/datasets").json()["data"]
    assert any(m["id"] == dataset_id for m in listed)

    resp = client.delete(f"/api/bi/datasets/{dataset_id}")
    assert resp.status_code == 200
    listed_after = client.get("/api/bi/datasets").json()["data"]
    assert all(m["id"] != dataset_id for m in listed_after)


def test_delete_missing_returns_404(client):
    resp = client.delete("/api/bi/datasets/ds_nope")
    assert resp.status_code == 404
    assert resp.json()["code"] == -1


# ─── 边界：文件类型 / 空文件 ─────────────────────────────────────
def test_reject_unsupported_extension(client):
    resp = _upload(client, "data.txt", b"hello")
    assert resp.status_code == 422
    assert "xlsx" in resp.json()["message"]


def test_reject_empty_file(client):
    resp = _upload(client, "empty.csv", b"")
    assert resp.status_code == 422


# ─── F-PV-02：分页 + 总行数（API-BI-04） ─────────────────────────
def test_preview_pagination_and_total(client):
    body = "名称,值\n" + "\n".join(f"r{i},{i}" for i in range(1, 121)) + "\n"
    did = _upload(client, "big.csv", _csv_bytes(body)).json()["data"]["id"]

    resp = client.post(
        f"/api/bi/datasets/{did}/preview", json={"page": 1, "pageSize": 50, "filters": []}
    )
    assert resp.status_code == 200
    data = resp.json()["data"]
    assert data["total"] == 120
    assert data["columns"] == ["名称", "值"]
    assert len(data["rows"]) == 50
    assert data["page"] == 1

    # 末页不足一页 → 余 20 行
    last = client.post(
        f"/api/bi/datasets/{did}/preview", json={"page": 3, "pageSize": 50, "filters": []}
    ).json()["data"]
    assert len(last["rows"]) == 20


# ─── F-PV-03：维度等值筛选 ───────────────────────────────────────
def test_preview_eq_filter(client):
    csv = "地区,销售额\n华东,1200\n华南,980\n华东,500\n"
    did = _upload(client, "s.csv", _csv_bytes(csv)).json()["data"]["id"]
    data = client.post(
        f"/api/bi/datasets/{did}/preview",
        json={"page": 1, "pageSize": 50, "filters": [{"field": "地区", "op": "eq", "value": "华东"}]},
    ).json()["data"]
    assert data["total"] == 2
    assert all(row[0] == "华东" for row in data["rows"])


# ─── F-PV-03：维度包含筛选 ───────────────────────────────────────
def test_preview_contains_filter(client):
    csv = "城市,人口\n北京,100\n南京,80\n东京,90\n上海,70\n"
    did = _upload(client, "c.csv", _csv_bytes(csv)).json()["data"]["id"]
    data = client.post(
        f"/api/bi/datasets/{did}/preview",
        json={"page": 1, "pageSize": 50, "filters": [{"field": "城市", "op": "contains", "value": "京"}]},
    ).json()["data"]
    assert data["total"] == 3


# ─── F-PV-03：数值范围筛选 ───────────────────────────────────────
def test_preview_numeric_range_filter(client):
    csv = "产品,销量\nA,10\nB,50\nC,100\n"
    did = _upload(client, "n.csv", _csv_bytes(csv)).json()["data"]["id"]
    data = client.post(
        f"/api/bi/datasets/{did}/preview",
        json={"page": 1, "pageSize": 50, "filters": [{"field": "销量", "op": "gte", "value": 50}]},
    ).json()["data"]
    assert data["total"] == 2


# ─── 边界：未知数据集 / 未知筛选列 ──────────────────────────────
def test_preview_unknown_dataset_404(client):
    resp = client.post(
        "/api/bi/datasets/ds_nope/preview", json={"page": 1, "pageSize": 50, "filters": []}
    )
    assert resp.status_code == 404


def test_preview_unknown_filter_field_ignored(client):
    csv = "a,b\n1,2\n3,4\n"
    did = _upload(client, "x.csv", _csv_bytes(csv)).json()["data"]["id"]
    data = client.post(
        f"/api/bi/datasets/{did}/preview",
        json={"page": 1, "pageSize": 50, "filters": [{"field": "zzz", "op": "eq", "value": "q"}]},
    ).json()["data"]
    assert data["total"] == 2


# ─── F-VZ-02 / API-BI-05：分组聚合（groupBy + sum） ──────────────
def test_aggregate_group_by_sum(client):
    csv = "地区,销售额\n华东,1200\n华南,980\n华东,500\n"
    did = _upload(client, "s.csv", _csv_bytes(csv)).json()["data"]["id"]
    resp = client.post(
        f"/api/bi/datasets/{did}/aggregate",
        json={
            "groupBy": ["地区"],
            "metrics": [{"field": "销售额", "agg": "sum", "alias": "总销售"}],
            "filters": [],
        },
    )
    assert resp.status_code == 200
    data = resp.json()["data"]
    assert data["columns"] == ["地区", "总销售"]
    result = {row[0]: row[1] for row in data["rows"]}
    assert result == {"华东": 1700, "华南": 980}


# ─── F-VZ-01：全量聚合（groupBy 为空）→ 单行总览 ────────────────
def test_aggregate_overview_no_groupby(client):
    csv = "地区,销售额\n华东,1200\n华南,980\n华东,500\n"
    did = _upload(client, "s.csv", _csv_bytes(csv)).json()["data"]["id"]
    data = client.post(
        f"/api/bi/datasets/{did}/aggregate",
        json={
            "groupBy": [],
            "metrics": [
                {"field": "销售额", "agg": "sum", "alias": "总销售"},
                {"field": "销售额", "agg": "avg", "alias": "均值"},
            ],
            "filters": [],
        },
    ).json()["data"]
    assert data["columns"] == ["总销售", "均值"]
    assert len(data["rows"]) == 1
    assert data["rows"][0][0] == 2680
    assert abs(data["rows"][0][1] - 2680 / 3) < 1e-6


# ─── API-BI-05：count_distinct + filters ─────────────────────────
def test_aggregate_count_distinct_with_filter(client):
    csv = "地区,客户\n华东,A\n华东,A\n华东,B\n华南,C\n"
    did = _upload(client, "cd.csv", _csv_bytes(csv)).json()["data"]["id"]
    data = client.post(
        f"/api/bi/datasets/{did}/aggregate",
        json={
            "groupBy": ["地区"],
            "metrics": [{"field": "客户", "agg": "count_distinct", "alias": "客户数"}],
            "filters": [{"field": "地区", "op": "eq", "value": "华东"}],
        },
    ).json()["data"]
    assert data["rows"] == [["华东", 2]]


# ─── 决策点 8：sum 跳过 NaN（缺失值不计入） ──────────────────────
def test_aggregate_skips_nan_in_sum(client):
    csv = "地区,销售额\n华东,100\n华东,\n华南,50\n"
    did = _upload(client, "nan.csv", _csv_bytes(csv)).json()["data"]["id"]
    data = client.post(
        f"/api/bi/datasets/{did}/aggregate",
        json={
            "groupBy": ["地区"],
            "metrics": [
                {"field": "销售额", "agg": "sum", "alias": "和"},
                {"field": "销售额", "agg": "count", "alias": "计数"},
            ],
            "filters": [],
        },
    ).json()["data"]
    result = {row[0]: (row[1], row[2]) for row in data["rows"]}
    assert result["华东"] == (100, 1)  # NaN 跳过：和=100，非空计数=1
    assert result["华南"] == (50, 1)


def test_aggregate_unknown_dataset_404(client):
    resp = client.post(
        "/api/bi/datasets/ds_nope/aggregate",
        json={"groupBy": [], "metrics": [{"field": "x", "agg": "sum", "alias": "x"}], "filters": []},
    )
    assert resp.status_code == 404


def test_aggregate_requires_at_least_one_metric(client):
    csv = "a,b\n1,2\n"
    did = _upload(client, "x.csv", _csv_bytes(csv)).json()["data"]["id"]
    resp = client.post(
        f"/api/bi/datasets/{did}/aggregate",
        json={"groupBy": ["a"], "metrics": [], "filters": []},
    )
    assert resp.status_code == 422
