"""BI 上传解析 — F-DS-01/02/03（PRD-bi §5.2 难点 2/3/10）。

职责（纯解析，不落盘）：
  · 读取上传的 xlsx / csv 字节流；
  · Excel 多级表头检测与拍平（"上级/子级"）；CSV 仅单级表头；
  · 类型推断（string/number/date/boolean）与默认角色（number→measure，其余→dimension）。
落盘由 bi_service 负责。解析失败统一抛 BiParseError → 路由层映射 422。
"""
from __future__ import annotations

import io
import re
import warnings
from typing import Any

import openpyxl
import pandas as pd

from app.services.bi_models import ColumnSchema

# object 列日期嗅探的最低命中率（避免把数字串误判为日期）
_DATE_SNIFF_MIN_RATIO = 0.9


class BiParseError(Exception):
    """解析失败（空文件 / 0 行 0 列 / 编码错乱等），路由层统一映射 422。"""


def parse(filename: str, content: bytes) -> tuple[list[ColumnSchema], pd.DataFrame]:
    """解析上传文件，返回 (列结构, 拍平后的 DataFrame)。

    filename 决定走 xlsx 还是 csv 分支（后缀已由路由层校验）。
    """
    lower = filename.lower()
    if lower.endswith(".xlsx"):
        df = _read_xlsx(content)
    else:
        df = _read_csv(content)

    if df.shape[0] == 0 or df.shape[1] == 0:
        raise BiParseError("文件解析失败，请检查文件内容")

    columns = [_infer_column(df[col]) for col in df.columns]
    return columns, df


# ─── CSV ─────────────────────────────────────────────────────────
def _read_csv(content: bytes) -> pd.DataFrame:
    """CSV：utf-8-sig 优先，失败回退 gbk（PRD 决策点 7）；单级表头。"""
    text: str | None = None
    for enc in ("utf-8-sig", "gbk"):
        try:
            text = content.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise BiParseError("文件解析失败，请检查文件内容")

    try:
        df = pd.read_csv(io.StringIO(text))
    except Exception as exc:  # noqa: BLE001 — pandas 异常种类多，统一兜底
        raise BiParseError("文件解析失败，请检查文件内容") from exc

    df.columns = _flatten_single(list(df.columns))
    return df


# ─── XLSX ────────────────────────────────────────────────────────
def _read_xlsx(content: bytes) -> pd.DataFrame:
    """XLSX：openpyxl 检测多级表头行数 N（合并单元格）→ 拍平为 "上级/子级"。"""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise BiParseError("文件解析失败，请检查文件内容") from exc

    ws = wb.active
    if ws is None or ws.max_row == 0 or ws.max_column == 0:
        raise BiParseError("文件解析失败，请检查文件内容")

    header_rows = _detect_header_rows(ws)
    names = _build_xlsx_headers(ws, header_rows)

    # 数据从表头之后开始；header=None + skiprows=N 让 pandas 仅推断数据区 dtype
    try:
        df = pd.read_excel(
            io.BytesIO(content), header=None, skiprows=header_rows, engine="openpyxl"
        )
    except Exception as exc:  # noqa: BLE001
        raise BiParseError("文件解析失败，请检查文件内容") from exc

    # 列数对齐：以解析出的数据列数为准（防表头/数据列数不一致）
    if df.shape[1] != len(names):
        names = (names + [f"列{i + 1}" for i in range(df.shape[1])])[: df.shape[1]]
    df.columns = _dedup(names)
    return df


def _detect_header_rows(ws: Any) -> int:
    """检测表头行数：首行存在横向合并单元格 → 判定 2 级表头，否则 1 级（PRD 难点 2）。"""
    for rng in ws.merged_cells.ranges:
        if rng.min_row == 1 and rng.max_col > rng.min_col:
            return 2
    return 1


def _build_xlsx_headers(ws: Any, header_rows: int) -> list[str]:
    """读取前 header_rows 行并对合并单元格做前向填充，逐列拼为 "上级/子级"。"""
    n_cols = ws.max_column
    # grid[r][c]：第 r 行第 c 列（0-based）的表头文本，合并单元格先填 None 再回填
    grid: list[list[str | None]] = [
        [_clean(ws.cell(row=r + 1, column=c + 1).value) for c in range(n_cols)]
        for r in range(header_rows)
    ]
    # 用合并区左上角值回填整个合并范围（仅表头区域内）
    for rng in ws.merged_cells.ranges:
        if rng.min_row > header_rows:
            continue
        anchor = _clean(ws.cell(row=rng.min_row, column=rng.min_col).value)
        for r in range(rng.min_row - 1, min(rng.max_row, header_rows)):
            for c in range(rng.min_col - 1, rng.max_col):
                grid[r][c] = anchor

    names: list[str] = []
    for c in range(n_cols):
        parts: list[str] = []
        for r in range(header_rows):
            part = grid[r][c]
            # 跳过空层级与纵向合并产生的重复层级（如 A1:A2 合并 → "地区/地区"）
            if part and (not parts or parts[-1] != part):
                parts.append(part)
        names.append("/".join(parts) if parts else f"列{c + 1}")
    return names


# ─── 表头规整与类型推断 ───────────────────────────────────────────
def _clean(value: Any) -> str | None:
    """单元格表头文本规整：None / 空白 → None。"""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _flatten_single(cols: list[Any]) -> list[str]:
    """单级表头：空列名（含 pandas 的 "Unnamed: N"）回退 列{i}。"""
    out: list[str] = []
    for i, col in enumerate(cols):
        text = "" if col is None else str(col).strip()
        if not text or text.startswith("Unnamed"):
            text = f"列{i + 1}"
        out.append(text)
    return _dedup(out)


def _dedup(names: list[str]) -> list[str]:
    """重名列追加后缀，避免 DataFrame 列名冲突。"""
    seen: dict[str, int] = {}
    out: list[str] = []
    for name in names:
        if name in seen:
            seen[name] += 1
            out.append(f"{name}_{seen[name]}")
        else:
            seen[name] = 0
            out.append(name)
    return out


def _infer_column(series: pd.Series) -> ColumnSchema:
    """单列类型推断 + 默认角色。number→measure，其余→dimension（PRD 难点 3）。"""
    name = str(series.name)
    inferred = _infer_type(series)
    # 高唯一率整数列疑似 ID（如订单号）→ 倾向维度，避免被当指标求和
    role: str = "measure" if inferred == "number" else "dimension"
    if inferred == "number" and _looks_like_id(series):
        role = "dimension"
    return ColumnSchema(name=name, type=inferred, role=role)  # type: ignore[arg-type]


def _infer_type(series: pd.Series) -> str:
    if pd.api.types.is_bool_dtype(series):
        return "boolean"
    if pd.api.types.is_datetime64_any_dtype(series):
        return "date"
    if pd.api.types.is_numeric_dtype(series):
        return "number"
    if _is_date_like(series):
        return "date"
    return "string"


def _is_date_like(series: pd.Series) -> bool:
    """object 列日期嗅探：抽样非空值 ≥90% 可解析为日期，且不是纯数字串。"""
    sample = series.dropna().astype(str).head(50)
    if sample.empty:
        return False
    if sample.str.fullmatch(r"-?\d+(\.\d+)?").all():
        return False  # 纯数字串不当日期
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")  # 抑制无固定格式的逐元素解析告警
        parsed = pd.to_datetime(sample, errors="coerce")
    return parsed.notna().mean() >= _DATE_SNIFF_MIN_RATIO


def _looks_like_id(series: pd.Series) -> bool:
    """疑似 ID 的整数列倾向维度。仅凭列名提示判定 —— 纯高唯一率信号会把正常
    数值指标（销售额/金额等天然各行不同）误判为 ID，故不采用（用户仍可改角色）。"""
    non_null = series.dropna()
    if non_null.empty:
        return False
    is_integer = pd.api.types.is_integer_dtype(series) or (non_null == non_null.round()).all()
    if not is_integer:
        return False
    name = str(series.name).lower()
    return bool(re.search(r"(id|编号|序号|订单号|no\.?$)", name))
